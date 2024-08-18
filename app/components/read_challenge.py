from pathlib import Path
from openpyxl import load_workbook

def read_challenge_data():
    challenge_path = Path.cwd() / 'resources' / 'challenge.xlsx'
    wb = load_workbook(challenge_path, read_only=True)
    ws = wb.active

    data = [row for row in ws.iter_rows(values_only=True,max_row=11,max_col=7,)]
    head = ('first_name','last_name','company_name','role_in_company', 'address','email','phone_number')
    people = [dict(zip(head, person)) for person in data[1:]]
    wb.close()
    
    return people